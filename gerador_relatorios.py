import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

def criar_relatorio_vendas():
    print("🚀 Iniciando o robô de relatórios do Excel...")
    
    # Cria uma nova planilha na memória do computador
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Relatório de Vendas"
    
    # Dados brutos que o robô vai inserir (Simulando um banco de dados de vendas)
    dados_vendas = [
        ["Produto", "Quantidade", "Preço Unitário (R$)", "Total Faturado (R$)"],
        ["Notebook Intel i7", 5, 4500.00, "=B2*C2"],
        ["Monitor UltraWide 29", 12, 1300.00, "=B3*C3"],
        ["Teclado Mecânico RGB", 25, 350.00, "=B4*C4"],
        ["Mouse Gamer Wireless", 40, 220.00, "=B5*C5"],
        ["Cadeira Ergonômica", 8, 1100.00, "=B6*C6"]
    ]
    
    print("✍️ Escrevendo as linhas e inserindo fórmulas matemáticas...")
    # Insere as linhas de dados na planilha uma por uma
    for linha in dados_vendas:
        sheet.append(linha)
        
    print("🎨 Aplicando design profissional na planilha...")
    # Estilização do cabeçalho (Linha 1)
    cor_azul_escuro = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    fonte_branca_negrito = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    alinhamento_centralizado = Alignment(horizontal="center", vertical="center")
    
    # Aplica os estilos apenas nas células do cabeçalho
    for celula in sheet[1]:
        celula.fill = cor_azul_escuro
        celula.font = fonte_branca_negrito
        celula.alignment = alinhamento_centralizado
        
    # Ajusta automaticamente a largura das colunas para o texto não ficar cortado
    for coluna in sheet.columns:
        maior_comprimento = max(len(str(celula.value or '')) for celula in coluna)
        letra_coluna = openpyxl.utils.get_column_letter(coluna[0].column)
        sheet.column_dimensions[letra_coluna].width = max(maior_comprimento + 4, 12)
        
    # Salva o arquivo final no computador
    nome_arquivo = "relatorio_vendas_automatico.xlsx"
    workbook.save(nome_arquivo)
    
    print(f"✨ Planilha criada com sucesso! Arquivo salvo como: [{nome_arquivo}]")

if __name__ == "__main__":
    criar_relatorio_vendas()
